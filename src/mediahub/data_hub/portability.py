"""data_hub.portability — CSV/XLSX in and out, deterministically (roadmap 1.13).

A club can bring its data in from a spreadsheet and take it back out again — a
clean **round-trip**. Like every parser in MediaHub this is deterministic and
*honest*: a cell that does not fit its column's kind is **flagged for review**,
never silently coerced to zero or dropped (the "flag ambiguous rows" rule). The
values it does accept are kept exactly.

* :func:`import_bytes` — sniff CSV vs XLSX, read the header into columns (kinds
  inferred, or matched to an existing table), and read each data row into cells
  stamped ``IMPORTED``. Anything it can't cleanly read becomes a flagged cell +
  a :class:`~mediahub.data_hub.models.DataWarning`.
* :func:`export_csv` / :func:`export_xlsx` — write a table back out (display
  values), so an exported sheet re-imports to the same logical data.

Only the standard library (``csv``) and the already-vendored ``openpyxl`` are
used; XLSX is read/written read-only-safe and honest-errors if openpyxl is
absent.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from typing import Optional

from mediahub.club_records.store import format_time_cs, parse_time_cs

from .models import DataCell, DataColumn, DataTable, DataWarning, Provenance

# Order matters: the first kind every non-empty cell fits wins. ``int`` before
# ``number`` (whole numbers stay whole); ``time`` after ``number`` so a plain
# decimal stays numeric and only a colon-bearing value ("1:05.32") reads as time.
_INFER_ORDER = ("int", "number", "date", "bool", "time", "text")

_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
_BOOL_TRUE = {"yes", "true", "y"}
_BOOL_FALSE = {"no", "false", "n"}


@dataclass
class ImportResult:
    table: Optional[DataTable]
    warnings: list[DataWarning]

    @property
    def ok(self) -> bool:
        return self.table is not None and self.table.row_count > 0

    def to_dict(self) -> dict:
        return {
            "ok": self.ok,
            "table": self.table.to_dict() if self.table else None,
            "warnings": [w.to_dict() for w in self.warnings],
        }


# ---------------------------------------------------------------------------
# Coercion + inference
# ---------------------------------------------------------------------------


def _coerce(text: str, col_type: str):
    """Return ``(value, ok)`` for ``text`` under ``col_type``.

    ``ok`` is False when the text does not fit the kind (so it gets flagged).
    The display always keeps the original text.
    """
    s = (text or "").strip()
    if s == "":
        return "", True  # an empty cell is allowed (a gap), never flagged
    if col_type == "text":
        return s, True
    if col_type == "int":
        try:
            return int(s.replace(",", "")), True
        except ValueError:
            return s, False
    if col_type == "number":
        try:
            return float(s.replace(",", "")), True
        except ValueError:
            return s, False
    if col_type == "time":
        cs = parse_time_cs(s)
        return (cs, True) if cs is not None else (s, False)
    if col_type == "date":
        return (s, True) if _DATE_RE.match(s) else (s, False)
    if col_type == "bool":
        low = s.lower()
        if low in _BOOL_TRUE:
            return True, True
        if low in _BOOL_FALSE:
            return False, True
        return s, False
    return s, True


def _infer_type(cells: list[str]) -> str:
    values = [c.strip() for c in cells if c and c.strip()]
    if not values:
        return "text"
    for kind in _INFER_ORDER:
        if all(_coerce(v, kind)[1] for v in values):
            return kind
    return "text"


def _display_for(value, col_type: str, raw: str) -> str:
    """How a coerced value is shown in the grid (and exported)."""
    if value in (None, ""):
        return raw if raw else ""
    if col_type == "time" and isinstance(value, int):
        return format_time_cs(value)
    if col_type == "bool" and isinstance(value, bool):
        return "Yes" if value else "No"
    return raw if raw else str(value)


# ---------------------------------------------------------------------------
# Reading rows from CSV / XLSX
# ---------------------------------------------------------------------------


def _read_csv_rows(text: str) -> list[list[str]]:
    if not text or not text.strip():
        return []
    sample = text[:2048]
    delim = ","
    try:
        delim = csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except csv.Error:
        if "\t" in sample and "," not in sample:
            delim = "\t"
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    return [[("" if c is None else str(c)) for c in r] for r in reader if r is not None]


def _read_xlsx_rows(data: bytes) -> list[list[str]]:
    try:
        import openpyxl
    except ImportError as exc:  # honest error — never a silent empty import
        raise ValueError(
            "Reading .xlsx needs the openpyxl library, which isn't available."
        ) from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — surface the real reason to the user
        raise ValueError(f"That .xlsx file could not be read ({exc}).") from exc
    rows: list[list[str]] = []
    try:
        ws = wb.active
        if ws is None:
            return []
        for raw in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in raw]
            if any(c.strip() for c in cells):
                rows.append(cells)
    finally:
        wb.close()
    return rows


def _detect_format(filename: str, data: bytes) -> str:
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return "xlsx"
    if name.endswith(".csv") or name.endswith(".txt") or name.endswith(".tsv"):
        return "csv"
    # XLSX is a zip; sniff the magic bytes if the name was unhelpful.
    if data[:2] == b"PK":
        return "xlsx"
    return "csv"


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------


def import_bytes(
    data: bytes,
    filename: str,
    *,
    title: str = "",
    existing_columns: Optional[list[DataColumn]] = None,
) -> ImportResult:
    """Parse uploaded CSV/XLSX bytes into a :class:`DataTable`.

    When ``existing_columns`` is given the import is matched to that table's
    columns (kinds and keys), and mismatching cells are flagged. Otherwise the
    header row names the columns and each kind is inferred from its values.
    """
    warnings: list[DataWarning] = []
    fmt = _detect_format(filename, data)
    try:
        if fmt == "xlsx":
            rows = _read_xlsx_rows(data)
        else:
            rows = _read_csv_rows(data.decode("utf-8-sig", errors="replace"))
    except ValueError as exc:
        warnings.append(DataWarning(0, str(exc), severity="error"))
        return ImportResult(None, warnings)

    if not rows:
        warnings.append(DataWarning(0, "The file is empty — nothing to import.", severity="error"))
        return ImportResult(None, warnings)

    header = [h.strip() for h in rows[0]]
    if not any(header):
        warnings.append(DataWarning(0, "The first row must name the columns.", severity="error"))
        return ImportResult(None, warnings)

    data_rows = rows[1:]

    # Decide columns: match an existing table, or infer from the data.
    if existing_columns:
        columns = list(existing_columns)
        col_keys = [c.key for c in columns]
        col_titles = [c.title for c in columns]
        # Map source header position → column index by title (case-insensitive).
        title_to_idx = {t.strip().lower(): i for i, t in enumerate(col_titles)}
        src_to_col: dict[int, int] = {}
        for src_i, h in enumerate(header):
            ci = title_to_idx.get(h.lower())
            if ci is not None:
                src_to_col[src_i] = ci
        if not src_to_col:
            warnings.append(
                DataWarning(
                    0, "None of the columns matched this table's columns.", severity="error"
                )
            )
            return ImportResult(None, warnings)
    else:
        col_titles = [h or f"Column {i + 1}" for i, h in enumerate(header)]
        col_keys = _unique_keys(col_titles)
        # Infer each column's kind from its data.
        columns = []
        for ci, (key, title) in enumerate(zip(col_keys, col_titles)):
            sample = [r[ci] if ci < len(r) else "" for r in data_rows]
            columns.append(
                DataColumn(key=key, title=title, type=_infer_type(sample), editable=True)
            )
        src_to_col = {i: i for i in range(len(columns))}

    # Build the rows.
    out_rows: list[dict] = []
    for ri, raw_row in enumerate(data_rows, start=1):
        if not any(str(c).strip() for c in raw_row):
            continue  # blank line — skip silently
        if len(raw_row) > len(header):
            warnings.append(
                DataWarning(ri, "Row had more cells than there are columns — extras ignored.")
            )
        cells: dict[str, DataCell] = {}
        for src_i, ci in src_to_col.items():
            col = columns[ci]
            raw = str(raw_row[src_i]).strip() if src_i < len(raw_row) else ""
            value, ok = _coerce(raw, col.type)
            display = _display_for(value, col.type, raw)
            flagged = (not ok) and raw != ""
            note = ""
            if flagged:
                note = f"'{raw}' isn't a valid {col.type} value."
                warnings.append(DataWarning(ri, f"{note} (column '{col.title}')", cell=raw))
            cells[col.key] = DataCell(
                value=(raw if flagged else value),
                display=display,
                provenance=Provenance.IMPORTED,
                source=filename or "import",
                note=note,
                flagged=flagged,
            )
        out_rows.append(cells)

    if not out_rows:
        warnings.append(DataWarning(0, "No data rows found under the header.", severity="error"))
        return ImportResult(None, warnings)

    table = DataTable(
        table_id="",  # assigned on save
        title=title.strip() or (filename.rsplit(".", 1)[0] if filename else "Imported table"),
        kind="org",
        columns=columns,
        rows=out_rows,
        editable=True,
        source=filename or "import",
        warnings=warnings,
    )
    return ImportResult(table, warnings)


def _unique_keys(titles: list[str]) -> list[str]:
    """Stable, unique snake_case keys from column titles."""
    keys: list[str] = []
    seen: dict[str, int] = {}
    for t in titles:
        base = re.sub(r"[^a-z0-9]+", "_", t.strip().lower()).strip("_") or "col"
        if base in seen:
            seen[base] += 1
            base = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        keys.append(base)
    return keys


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


def _row_display(table: DataTable, row: dict, key: str) -> str:
    cell = row.get(key)
    if cell is None:
        return ""
    if not isinstance(cell, DataCell):
        cell = DataCell.from_dict(cell)
    return cell.display if cell.display else ("" if cell.value is None else str(cell.value))


def export_csv(table: DataTable) -> str:
    """Render a table to CSV text (header titles + display values)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([c.title for c in table.columns])
    for row in table.rows:
        writer.writerow([_row_display(table, row, c.key) for c in table.columns])
    return buf.getvalue()


def export_xlsx(table: DataTable) -> bytes:
    """Render a table to XLSX bytes (header titles + display values)."""
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError(
            "Writing .xlsx needs the openpyxl library, which isn't available."
        ) from exc
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (table.title or "Sheet")[:31]
    ws.append([c.title for c in table.columns])
    for row in table.rows:
        ws.append([_row_display(table, row, c.key) for c in table.columns])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


__all__ = [
    "ImportResult",
    "import_bytes",
    "export_csv",
    "export_xlsx",
]
